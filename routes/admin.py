from flask import Blueprint, render_template, request, jsonify, session, redirect, url_for, flash, Response
from db import get_db_connection
from extensions import send_email
import uuid, math, io
from datetime import datetime, timedelta
import cloudinary.uploader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

admin_bp = Blueprint('admin', __name__)

def _wajib_login_admin():
    if 'user_id' not in session or (session.get('role') or '').lower() != 'kepala':
        session.clear()
        return redirect(url_for('admin.login_admin'))
    return None

def _ke_menit(t):
    if t is None:
        return None
    if hasattr(t, 'total_seconds'): 
        return int(t.total_seconds() // 60)
    if hasattr(t, 'hour'):  
        return t.hour * 60 + t.minute
    return None


@admin_bp.app_template_filter('rupiah')
def format_rupiah(value):
    try:
        value = int(value)
        return "Rp{:,.0f}".format(value).replace(",", ".")
    except (ValueError, TypeError):
        return "Rp0"


# --- HELPER: Format tanggal ke Bahasa Indonesia (dilakukan di Python, bukan SQL,
# supaya tidak tergantung cara driver database menangani karakter '%') ---
_NAMA_BULAN_ID = {
    1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'Mei', 6: 'Jun',
    7: 'Jul', 8: 'Agu', 9: 'Sep', 10: 'Okt', 11: 'Nov', 12: 'Des'
}

def format_tanggal_indo(dt):
    if not dt:
        return '-'
    try:
        return f"{dt.day:02d} {_NAMA_BULAN_ID[dt.month]} {dt.year}, {dt.hour:02d}:{dt.minute:02d}"
    except (AttributeError, KeyError):
        return str(dt)

@admin_bp.route('/dashboard_admin')
def dashboard_admin():
    print("ISI SESSION:", dict(session))
    cek_akses = _wajib_login_admin()
    if cek_akses:
        return cek_akses
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 1. Hitung Total Murid Sebenarnya (Dari tabel anak)
        cursor.execute("SELECT COUNT(*) as total FROM anak")
        murid = cursor.fetchone()['total'] or 0
        
        # 2. Hitung Total Pengajar Sebenarnya (Berdasarkan ID Role 'R02')
        cursor.execute("SELECT COUNT(*) as total FROM users WHERE role_id_role = 'R02'")
        pengajar = cursor.fetchone()['total'] or 0 

        # 3. Nilai Riil Kelas Aktif (Tabel kelas sudah ada di database)
        cursor.execute("SELECT COUNT(*) as total FROM kelas WHERE status_kelas = 'Aktif'")
        kelas_aktif = cursor.fetchone()['total'] or 0 
        
        # 4. Hitung Total Orang Tua/Keluarga yang masih 'unverified' (Pending)
        cursor.execute("""
            SELECT COUNT(*) as total 
            FROM users u
            JOIN role r ON u.role_id_role = r.id_role
            WHERE r.nama_role = 'Murid' AND u.status_akun = 'unverified'
        """)
        pending = cursor.fetchone()['total'] or 0 

        # 5. Kelas Favorit: kelas dengan jumlah murid aktif terbanyak
        cursor.execute("""
            SELECT 
                k.nama_kelas,
                k.tingkat,
                k.kapasitas_maksimal,
                COUNT(p.id_pendaftaran) AS jumlah_siswa
            FROM kelas k
            LEFT JOIN pendaftaran p ON k.id_kelas = p.id_kelas AND p.status_pendaftaran = 'Aktif'
            GROUP BY k.id_kelas
            ORDER BY jumlah_siswa DESC
            LIMIT 5
        """)
        kelas_favorit = cursor.fetchall()

        # 6. Breakdown Total Murid per jenjang (SD/SMP/SMA) dari kolom anak.kelas.
        # Pakai LIKE supaya tetap kena walau formatnya bervariasi (mis. "SD",
        # "Kelas 5 SD", dst) -- kalau di database kamu formatnya ternyata beda,
        # kasih tahu saya biar pattern-nya disesuaikan.
        cursor.execute("""
            SELECT
                SUM(CASE WHEN kelas LIKE '%SD%' THEN 1 ELSE 0 END) AS total_sd,
                SUM(CASE WHEN kelas LIKE '%SMP%' THEN 1 ELSE 0 END) AS total_smp,
                SUM(CASE WHEN kelas LIKE '%SMA%' THEN 1 ELSE 0 END) AS total_sma
            FROM anak
        """)
        jenjang_row = cursor.fetchone() or {}
        murid_jenjang = {
            'sd': jenjang_row.get('total_sd') or 0,
            'smp': jenjang_row.get('total_smp') or 0,
            'sma': jenjang_row.get('total_sma') or 0,
        }
        murid_jenjang['max'] = max(murid_jenjang['sd'], murid_jenjang['smp'], murid_jenjang['sma'], 1)

        return render_template('admin/dashboard_admin.html', 
                               murid=murid, 
                               pengajar=pengajar,
                               kelas_aktif=kelas_aktif,
                               pending=pending,
                               kelas_favorit=kelas_favorit,
                               murid_jenjang=murid_jenjang,
                               username=session.get('username', 'Admin Dennis'))
                               
    except Exception as e:
        print(f"\n[ERROR DASHBOARD]: {e}\n")
        return render_template('admin/dashboard_admin.html', murid=0, pengajar=0, kelas_aktif=0, pending=0, kelas_favorit=[], 
                               murid_jenjang={'sd': 0, 'smp': 0, 'sma': 0, 'max': 1}, username="Admin")
    finally:
        cursor.close()
        conn.close()
# Tugas: Menampilkan wujud halaman web (UI)
@admin_bp.route('/login_admin')
def login_admin():
    return render_template('admin/login_admin.html')

@admin_bp.route('/login_admin_action', methods=['POST'])
def login_admin_action():
    data = request.get_json(force=True)
    email = data.get('email')
    password = data.get('password')

    if not email or not password:
        return jsonify({"message": "Email dan kata sandi harus diisi!"}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 1. PERBAIKAN: Menggunakan u.role_id_role sesuai struktur tabel Anda
        cursor.execute("""
            SELECT u.*, r.nama_role 
            FROM users u
            JOIN role r ON u.role_id_role = r.id_role
            WHERE u.email = %s AND r.nama_role = 'Kepala'
        """, (email,))
        
        admin_data = cursor.fetchone()

        if admin_data:
            # 2. PERBAIKAN: Karena password Admin ('12345678') di DB bukan hash scrypt, 
            # kita harus menggunakan pengecekan sama dengan (==) biasa.
            if admin_data['password'] == password:
                
                # Buat sesi login
                session['user_id'] = admin_data['id_users'] 
                session['role'] = admin_data['nama_role']
                
                # Bonus: Simpan username agar bisa ditampilkan di navbar dashboard nanti!
                session['username'] = admin_data['username'] 

                return jsonify({
                    "message": "Login berhasil!", 
                    "redirect": "/dashboard_admin"
                }), 200
            else:
                return jsonify({"message": "Kata sandi salah!"}), 401
        else:
            return jsonify({"message": "Akun admin tidak ditemukan atau email salah!"}), 404

    except Exception as e:
        print(f"Error saat login admin: {e}")
        return jsonify({"message": "Terjadi kesalahan pada server."}), 500
    finally:
        if cursor.with_rows:
            cursor.fetchall()
        cursor.close()
        conn.close()

@admin_bp.route('/logout_admin')
def logout_admin():
    session.clear()
    return redirect(url_for('admin.login_admin'))

@admin_bp.route('/manajemen_kelas_admin')
def manajemen_kelas_admin():
    # 1. Proteksi Akses Admin Kepala
    cek_akses = _wajib_login_admin()
    if cek_akses:
        return cek_akses
        
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 2. Hitung statistik ringkas untuk kartu bagian atas halaman
        cursor.execute("SELECT COUNT(*) as total FROM kelas")
        total_kelas = cursor.fetchone()['total'] or 0
        
        cursor.execute("SELECT COUNT(*) as total FROM kelas WHERE status_kelas = 'Aktif'")
        kelas_aktif = cursor.fetchone()['total'] or 0
        
        cursor.execute("SELECT SUM(kapasitas_maksimal) as total_slot FROM kelas")
        fetch_slot = cursor.fetchone()
        total_slot = fetch_slot['total_slot'] if fetch_slot and fetch_slot['total_slot'] is not None else 0
        
        # 3. Query Utama: Mengambil data persis sesuai kolom di tec_english.sql terbaru
        query_kelas = """
            SELECT 
                k.id_kelas,
                k.nama_kelas,
                k.hari_jadwal,
                k.jam_mulai,
                k.jam_selesai,
                k.kapasitas_maksimal,
                k.harga,
                k.status_kelas,
                u.username AS nama_tutor,
                (SELECT COUNT(*) FROM pendaftaran p WHERE p.id_kelas = k.id_kelas) AS jumlah_siswa
            FROM kelas k
            LEFT JOIN users u ON k.id_pengajar = u.id_users
            ORDER BY k.created_at DESC
        """
        cursor.execute(query_kelas)
        daftar_kelas = cursor.fetchall()
        
        # 4. Konversi Data TIME & DECIMAL agar aman dikirim ke Jinja2 HTML
        for kelas in daftar_kelas:
            # Format Jam Mulai
            if kelas['jam_mulai']:
                if hasattr(kelas['jam_mulai'], 'total_seconds'):  # jika bertipe timedelta
                    total_sec = int(kelas['jam_mulai'].total_seconds())
                    kelas['jam_mulai'] = f"{total_sec // 3600:02d}:{(total_sec % 3600) // 60:02d}"
                elif hasattr(kelas['jam_mulai'], 'strftime'):  # jika bertipe time object
                    kelas['jam_mulai'] = kelas['jam_mulai'].strftime('%H:%M')
                else:
                    kelas['jam_mulai'] = str(kelas['jam_mulai'])[:5]

            # Format Jam Selesai
            if kelas['jam_selesai']:
                if hasattr(kelas['jam_selesai'], 'total_seconds'):
                    total_sec = int(kelas['jam_selesai'].total_seconds())
                    kelas['jam_selesai'] = f"{total_sec // 3600:02d}:{(total_sec % 3600) // 60:02d}"
                elif hasattr(kelas['jam_selesai'], 'strftime'):
                    kelas['jam_selesai'] = kelas['jam_selesai'].strftime('%H:%M')
                else:
                    kelas['jam_selesai'] = str(kelas['jam_selesai'])[:5]
            
            # Pastikan harga diubah ke integer/float murni (bukan objek Decimal)
            kelas['harga'] = int(kelas['harga']) if kelas['harga'] else 0
        
        return render_template(
            'admin/manajemen_kelas.html', 
            daftar_kelas=daftar_kelas,
            total_kelas=total_kelas,
            kelas_aktif=kelas_aktif,
            total_slot=total_slot,
            username=session.get('username', 'Admin')  
        )
        
    except Exception as e:
        print(f"\n[ERROR MANAJEMEN KELAS]: {e}\n")
        flash("Gagal memuat data kelas dari database.", "error")
        return render_template('admin/manajemen_kelas.html', daftar_kelas=[], total_kelas=0, kelas_aktif=0, total_slot=0,
                               username=session.get('username', 'Admin'))
        
    finally:
        cursor.close()
        conn.close()

@admin_bp.route('/buat_kelas_baru')
def buat_kelas_baru():
    # 1. Proteksi Sesi Admin (Hanya Kepala/Admin yang bisa akses)
    cek_akses = _wajib_login_admin()
    if cek_akses:
        return cek_akses
        
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 2. Ambil daftar user yang memiliki role Pengajar (R02) untuk dikirim ke dropdown HTML
        cursor.execute("SELECT id_users, username, email, foto_profil FROM users WHERE role_id_role = 'R02'")
        daftar_tutor = cursor.fetchall()
        
        return render_template('admin/tambah_kelas.html', daftar_tutor=daftar_tutor)
    except Exception as e:
        print(f"Error memuat form tambah kelas: {e}")
        return render_template('admin/tambah_kelas.html', daftar_tutor=[])
    finally:
        cursor.close()
        conn.close()


@admin_bp.route('/simpan_kelas_baru', methods=['POST'])
def simpan_kelas_baru():
    # Proteksi Sesi
    cek_akses = _wajib_login_admin()
    if cek_akses:
        return cek_akses
        
    # 1. Ambil data dari input form HTML
    nama_kelas = request.form.get('nama_kelas')
    tingkat = request.form.get('tingkat')
    id_pengajar = request.form.get('id_pengajar')
    hari_jadwal = request.form.get('hari_jadwal')
    jam_mulai = request.form.get('jam_mulai')
    durasi_menit = request.form.get('durasi_menit')  # total durasi kelas dalam menit (dari jam + menit di form)
    kapasitas_maksimal = request.form.get('kapasitas_maksimal')
    harga = request.form.get('harga')
    deskripsi = request.form.get('deskripsi')
    tanggal_mulai = request.form.get('tanggal_mulai')
    jumlah_sesi = request.form.get('jumlah_sesi')

    # 2. Validasi kelengkapan data
    if not all([nama_kelas, tingkat, id_pengajar, hari_jadwal, jam_mulai, durasi_menit, 
                kapasitas_maksimal, harga, deskripsi, tanggal_mulai, jumlah_sesi]):
        flash('Semua kolom formulir wajib diisi!', 'error')
        return redirect(url_for('admin.buat_kelas_baru'))

    try:
        durasi_menit = int(durasi_menit)
        jumlah_sesi = int(jumlah_sesi)
        if durasi_menit <= 0 or jumlah_sesi <= 0:
            raise ValueError
    except ValueError:
        flash('Durasi kelas dan jumlah sesi harus berupa angka yang valid!', 'error')
        return redirect(url_for('admin.buat_kelas_baru'))

    # 3. Hitung jam selesai otomatis dari jam mulai + durasi (dihitung ulang di server,
    # tidak cuma percaya hasil kalkulasi JS di browser)
    waktu_mulai = datetime.strptime(jam_mulai, '%H:%M')
    waktu_selesai = waktu_mulai + timedelta(minutes=durasi_menit)
    jam_selesai = waktu_selesai.strftime('%H:%M')

    # 4. Hitung tanggal tiap sesi: 1 sesi = 1 minggu, jadi sesi ke-n jatuh (n-1) minggu
    # setelah tanggal mulai. Tanggal berakhir kelas otomatis = tanggal sesi terakhir.
    try:
        tanggal_mulai_obj = datetime.strptime(tanggal_mulai, '%Y-%m-%d').date()
    except ValueError:
        flash('Format tanggal mulai tidak valid!', 'error')
        return redirect(url_for('admin.buat_kelas_baru'))

    tanggal_sesi_list = [tanggal_mulai_obj + timedelta(weeks=i) for i in range(jumlah_sesi)]
    tanggal_berakhir_obj = tanggal_sesi_list[-1]

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 5. VALIDASI BENTROK JADWAL PENGAJAR
        # Tolak pembuatan kelas kalau pengajar yang dipilih sudah punya kelas lain
        # (kecuali yang berstatus 'Selesai') di hari yang sama dengan jam yang
        # tumpang tindih -- satu pengajar tidak mungkin mengajar 2 kelas sekaligus.
        cursor.execute("""
            SELECT nama_kelas, hari_jadwal, jam_mulai, jam_selesai
            FROM kelas
            WHERE id_pengajar = %s
              AND hari_jadwal = %s
              AND status_kelas != 'Selesai'
        """, (id_pengajar, hari_jadwal))
        kelas_pengajar_di_hari_sama = cursor.fetchall()

        menit_mulai_baru = _ke_menit(waktu_mulai.time())
        menit_selesai_baru = _ke_menit(waktu_selesai.time())

        for kl in kelas_pengajar_di_hari_sama:
            menit_mulai_lama = _ke_menit(kl['jam_mulai'])
            menit_selesai_lama = _ke_menit(kl['jam_selesai'])

            # Dua rentang waktu tumpang tindih kalau: mulai_baru < selesai_lama DAN mulai_lama < selesai_baru
            if menit_mulai_baru < menit_selesai_lama and menit_mulai_lama < menit_selesai_baru:
                jam_lama_mulai = str(kl['jam_mulai'])[:5] if kl['jam_mulai'] is not None else '-'
                jam_lama_selesai = str(kl['jam_selesai'])[:5] if kl['jam_selesai'] is not None else '-'
                flash(
                    f"Jadwal bentrok! Pengajar ini sudah mengajar kelas \"{kl['nama_kelas']}\" "
                    f"pada hari {kl['hari_jadwal']} jam {jam_lama_mulai}-{jam_lama_selesai} WIB. "
                    f"Tidak bisa membuat kelas baru karena jadwalnya bertabrakan.",
                    'error'
                )
                return redirect(url_for('admin.buat_kelas_baru'))

        # 6. Generate ID kelas secara unik
        id_kelas_baru = f"KLS-{str(uuid.uuid4().hex[:6]).upper()}"

        # 7. Upload gambar sampul kelas ke Cloudinary (opsional -- boleh kosong)
        url_gambar_kelas = None
        file_gambar = request.files.get('gambar_kelas')
        if file_gambar and file_gambar.filename:
            try:
                hasil_upload = cloudinary.uploader.upload(
                    file_gambar,
                    folder="tec_portal/sampul_kelas",
                    resource_type="image"
                )
                url_gambar_kelas = hasil_upload.get('secure_url')
            except Exception as e:
                print(f"\n[ERROR UPLOAD SAMPUL KELAS]: {e}\n")
                flash('Gagal mengunggah gambar sampul kelas. Kelas belum disimpan, coba lagi.', 'error')
                return redirect(url_for('admin.buat_kelas_baru'))

        # 8. Eksekusi query INSERT ke tabel kelas
        query = """
            INSERT INTO kelas (
                id_kelas, nama_kelas, tingkat, id_pengajar, 
                hari_jadwal, jam_mulai, jam_selesai, 
                kapasitas_maksimal, harga, deskripsi, gambar_kelas,
                jumlah_sesi, tanggal_mulai, tanggal_berakhir, status_kelas
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'Aktif')
        """
        values = (
            id_kelas_baru, nama_kelas, tingkat, id_pengajar, 
            hari_jadwal, jam_mulai, jam_selesai, 
            kapasitas_maksimal, harga, deskripsi, url_gambar_kelas,
            jumlah_sesi, tanggal_mulai_obj, tanggal_berakhir_obj
        )
        cursor.execute(query, values)

        # 9. Generate otomatis baris-baris sesi di tabel sesi_kelas sebanyak jumlah_sesi,
        # masing-masing berjarak 1 minggu dari sesi sebelumnya
        for i, tanggal_sesi in enumerate(tanggal_sesi_list, start=1):
            cursor.execute("""
                INSERT INTO sesi_kelas (id_kelas, sesi_ke, tanggal)
                VALUES (%s, %s, %s)
            """, (id_kelas_baru, i, tanggal_sesi))

        conn.commit() 

        flash(
            f'Kelas "{nama_kelas}" berhasil dibuat dengan {jumlah_sesi} sesi '
            f'({tanggal_mulai_obj.strftime("%d %b %Y")} - {tanggal_berakhir_obj.strftime("%d %b %Y")})!',
            'success'
        )
        return redirect(url_for('admin.manajemen_kelas_admin'))

    except Exception as e:
        conn.rollback() 
        print(f"\n[ERROR SIMPAN KELAS]: {e}\n")
        flash('Terjadi kegagalan sistem saat menyimpan data kelas baru.', 'error')
        return redirect(url_for('admin.buat_kelas_baru'))
        
    finally:
        cursor.close()
        conn.close()


@admin_bp.route('/edit_kelas/<id_kelas>')
def edit_kelas(id_kelas):
    # 1. Proteksi Sesi Admin
    cek_akses = _wajib_login_admin()
    if cek_akses:
        return cek_akses

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 2. Ambil data kelas + jumlah siswa yang sudah terdaftar (dipakai untuk
        # menentukan field mana yang boleh diedit dan mana yang harus dikunci)
        cursor.execute("""
            SELECT k.*,
                (SELECT COUNT(*) FROM pendaftaran p WHERE p.id_kelas = k.id_kelas) AS jumlah_siswa
            FROM kelas k
            WHERE k.id_kelas = %s
        """, (id_kelas,))
        kelas = cursor.fetchone()

        if not kelas:
            flash('Kelas tidak ditemukan.', 'error')
            return redirect(url_for('admin.manajemen_kelas_admin'))

        # 3. Format jam (TIME/timedelta -> "HH:MM") supaya aman dipakai sebagai
        # value input time/date di form, sama seperti di manajemen_kelas_admin
        for field in ('jam_mulai', 'jam_selesai'):
            nilai = kelas.get(field)
            if nilai:
                if hasattr(nilai, 'total_seconds'):
                    total_sec = int(nilai.total_seconds())
                    kelas[field] = f"{total_sec // 3600:02d}:{(total_sec % 3600) // 60:02d}"
                elif hasattr(nilai, 'strftime'):
                    kelas[field] = nilai.strftime('%H:%M')

        kelas['harga'] = int(kelas['harga']) if kelas['harga'] else 0
        kelas['tanggal_mulai_str'] = kelas['tanggal_mulai'].strftime('%Y-%m-%d') if kelas.get('tanggal_mulai') else ''

        # 4. Hitung ulang durasi kelas (jam + menit) dari selisih jam_mulai & jam_selesai,
        # supaya dropdown durasi di form ter-preselect sesuai data lama
        durasi_jam, durasi_menit = 1, 0
        if kelas['jam_mulai'] and kelas['jam_selesai']:
            h1, m1 = map(int, kelas['jam_mulai'].split(':'))
            h2, m2 = map(int, kelas['jam_selesai'].split(':'))
            total_menit = (h2 * 60 + m2) - (h1 * 60 + m1)
            if total_menit < 0:
                total_menit += 24 * 60
            durasi_jam, durasi_menit = divmod(total_menit, 60)

        # 5. Daftar tutor untuk dropdown "Pilih Pengajar"
        cursor.execute("SELECT id_users, username, email, foto_profil FROM users WHERE role_id_role = 'R02'")
        daftar_tutor = cursor.fetchall()

        # 6. Daftar siswa yang sudah terdaftar di kelas ini, untuk widget referensi admin
        cursor.execute("""
            SELECT a.nama_lengkap, a.nama_panggilan, p.status_pendaftaran, p.tanggal_daftar
            FROM pendaftaran p
            JOIN anak a ON p.id_anak = a.id_anak
            WHERE p.id_kelas = %s
            ORDER BY p.tanggal_daftar ASC
        """, (id_kelas,))
        daftar_siswa = cursor.fetchall()

        return render_template('admin/edit_kelas.html',
                               kelas=kelas,
                               daftar_tutor=daftar_tutor,
                               daftar_siswa=daftar_siswa,
                               kelas_kosong=(kelas['jumlah_siswa'] == 0),
                               durasi_jam=durasi_jam,
                               durasi_menit=durasi_menit)

    except Exception as e:
        print(f"\n[ERROR EDIT KELAS]: {e}\n")
        flash('Gagal memuat data kelas untuk diedit.', 'error')
        return redirect(url_for('admin.manajemen_kelas_admin'))
    finally:
        cursor.close()
        conn.close()


@admin_bp.route('/update_kelas/<id_kelas>', methods=['POST'])
def update_kelas(id_kelas):
    # 1. Proteksi Sesi Admin
    cek_akses = _wajib_login_admin()
    if cek_akses:
        return cek_akses

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        
        # 2. Ambil ulang data kelas + jumlah siswa LANGSUNG DARI DATABASE (bukan
        # percaya nilai yang (mungkin) ikut terkirim dari form). Ini kunci utama
        # supaya field yang seharusnya terkunci tetap aman walau ada yang coba
        # mengakalinya lewat DevTools/request manual.
        cursor.execute("""
            SELECT k.*,
                (SELECT COUNT(*) FROM pendaftaran p WHERE p.id_kelas = k.id_kelas) AS jumlah_siswa
            FROM kelas k
            WHERE k.id_kelas = %s
        """, (id_kelas,))
        kelas_lama = cursor.fetchone()

        if not kelas_lama:
            flash('Kelas tidak ditemukan.', 'error')
            return redirect(url_for('admin.manajemen_kelas_admin'))

        kelas_kosong = kelas_lama['jumlah_siswa'] == 0

        # 3. Field yang SELALU boleh diedit (kosong ataupun sudah ada siswa)
        nama_kelas = request.form.get('nama_kelas')
        # `tingkat` SENGAJA tidak diambil dari form -- field ini dikunci permanen
        # di halaman edit (tingkat kelas tidak boleh berubah setelah dibuat),
        # jadi nilai lama dari database yang selalu dipakai apapun yang terkirim.
        tingkat = kelas_lama['tingkat']
        id_pengajar = request.form.get('id_pengajar')
        deskripsi = request.form.get('deskripsi')
        kapasitas_maksimal = request.form.get('kapasitas_maksimal')
        status_kelas = request.form.get('status_kelas')

        if not all([nama_kelas, id_pengajar, deskripsi, kapasitas_maksimal, status_kelas]):
            flash('Nama kelas, pengajar, deskripsi, kapasitas, dan status wajib diisi!', 'error')
            return redirect(url_for('admin.edit_kelas', id_kelas=id_kelas))

        if status_kelas not in ('Aktif', 'Nonaktif', 'Selesai', 'Penuh'):
            flash('Status kelas tidak valid.', 'error')
            return redirect(url_for('admin.edit_kelas', id_kelas=id_kelas))

        try:
            kapasitas_maksimal = int(kapasitas_maksimal)
            if kapasitas_maksimal < 1:
                raise ValueError
        except ValueError:
            flash('Kapasitas maksimal harus berupa angka yang valid!', 'error')
            return redirect(url_for('admin.edit_kelas', id_kelas=id_kelas))

        # Kapasitas tidak boleh diturunkan sampai di bawah jumlah siswa yang
        # sudah terdaftar -- ini dicek terhadap angka ASLI dari database, bukan
        # dari form, supaya tidak bisa dibobol.
        if kapasitas_maksimal < kelas_lama['jumlah_siswa']:
            flash(
                f'Kapasitas maksimal tidak boleh lebih kecil dari jumlah siswa yang sudah '
                f'terdaftar saat ini ({kelas_lama["jumlah_siswa"]} siswa).',
                'error'
            )
            return redirect(url_for('admin.edit_kelas', id_kelas=id_kelas))

        # 4. Field yang HANYA boleh diedit selama kelas masih kosong (belum ada
        # siswa) -- harga, jadwal/jam, tanggal mulai & jumlah sesi. Kalau kelas
        # sudah ada siswa, field ini diabaikan sepenuhnya dan nilai LAMA dari
        # database yang tetap dipakai, apapun yang terkirim dari form.
        regenerasi_sesi = False
        tanggal_sesi_list = []

        if kelas_kosong:
            hari_jadwal = request.form.get('hari_jadwal')
            jam_mulai = request.form.get('jam_mulai')
            durasi_menit_input = request.form.get('durasi_menit')
            tanggal_mulai = request.form.get('tanggal_mulai')
            jumlah_sesi = request.form.get('jumlah_sesi')
            harga = request.form.get('harga')

            if not all([hari_jadwal, jam_mulai, durasi_menit_input, tanggal_mulai, jumlah_sesi, harga]):
                flash('Semua kolom jadwal, tanggal, dan harga wajib diisi!', 'error')
                return redirect(url_for('admin.edit_kelas', id_kelas=id_kelas))

            try:
                durasi_menit_input = int(durasi_menit_input)
                jumlah_sesi = int(jumlah_sesi)
                if durasi_menit_input <= 0 or jumlah_sesi <= 0:
                    raise ValueError
            except ValueError:
                flash('Durasi kelas dan jumlah sesi harus berupa angka yang valid!', 'error')
                return redirect(url_for('admin.edit_kelas', id_kelas=id_kelas))

            waktu_mulai = datetime.strptime(jam_mulai, '%H:%M')
            waktu_selesai = waktu_mulai + timedelta(minutes=durasi_menit_input)
            jam_selesai = waktu_selesai.strftime('%H:%M')

            try:
                tanggal_mulai_obj = datetime.strptime(tanggal_mulai, '%Y-%m-%d').date()
            except ValueError:
                flash('Format tanggal mulai tidak valid!', 'error')
                return redirect(url_for('admin.edit_kelas', id_kelas=id_kelas))

            tanggal_sesi_list = [tanggal_mulai_obj + timedelta(weeks=i) for i in range(jumlah_sesi)]
            tanggal_berakhir_obj = tanggal_sesi_list[-1]
            regenerasi_sesi = True
        else:
            # Kelas sudah jalan -> pakai nilai lama apa adanya, field terkait
            # dianggap terkunci walau form (seharusnya) sudah men-disable-nya.
            hari_jadwal = kelas_lama['hari_jadwal']
            jam_mulai = kelas_lama['jam_mulai']
            jam_selesai = kelas_lama['jam_selesai']
            tanggal_mulai_obj = kelas_lama['tanggal_mulai']
            tanggal_berakhir_obj = kelas_lama['tanggal_berakhir']
            jumlah_sesi = kelas_lama['jumlah_sesi']
            harga = kelas_lama['harga']

            if hasattr(jam_mulai, 'total_seconds'):
                jam_mulai = (datetime.min + jam_mulai).time()
            if hasattr(jam_selesai, 'total_seconds'):
                jam_selesai = (datetime.min + jam_selesai).time()

        # 5. Simpan perubahan ke tabel kelas
        query = """
            UPDATE kelas SET
                nama_kelas = %s, tingkat = %s, id_pengajar = %s, deskripsi = %s,
                kapasitas_maksimal = %s, status_kelas = %s,
                hari_jadwal = %s, jam_mulai = %s, jam_selesai = %s,
                tanggal_mulai = %s, tanggal_berakhir = %s, jumlah_sesi = %s, harga = %s
            WHERE id_kelas = %s
        """
        cursor.execute(query, (
            nama_kelas, tingkat, id_pengajar, deskripsi,
            kapasitas_maksimal, status_kelas,
            hari_jadwal, jam_mulai, jam_selesai,
            tanggal_mulai_obj, tanggal_berakhir_obj, jumlah_sesi, harga,
            id_kelas
        ))

        # 6. Kalau jadwal/tanggal/jumlah sesi ikut berubah (hanya mungkin saat
        # kelas masih kosong), generate ulang baris sesi_kelas dari awal. Aman
        # dihapus karena kelas kosong = belum ada siswa = belum ada presensi.
        if regenerasi_sesi:
            cursor.execute("DELETE FROM sesi_kelas WHERE id_kelas = %s", (id_kelas,))
            for i, tanggal_sesi in enumerate(tanggal_sesi_list, start=1):
                cursor.execute("""
                    INSERT INTO sesi_kelas (id_kelas, sesi_ke, tanggal)
                    VALUES (%s, %s, %s)
                """, (id_kelas, i, tanggal_sesi))

        conn.commit()
        flash(f'Kelas "{nama_kelas}" berhasil diperbarui!', 'success')
        return redirect(url_for('admin.manajemen_kelas_admin'))

    except Exception as e:
        conn.rollback()
        print(f"\n[ERROR UPDATE KELAS]: {e}\n")
        flash('Terjadi kegagalan sistem saat memperbarui data kelas.', 'error')
        return redirect(url_for('admin.edit_kelas', id_kelas=id_kelas))
    finally:
        cursor.close()
        conn.close()


@admin_bp.route('/hapus_kelas/<id_kelas>', methods=['POST'])
def hapus_kelas(id_kelas):
    # 1. Proteksi Sesi Admin
    cek_akses = _wajib_login_admin()
    if cek_akses:
        return cek_akses

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 2. Cek ulang langsung dari database: kelas ini benar-benar belum
        # punya siswa sama sekali? (jangan percaya kondisi apapun dari sisi
        # frontend/tombol yang mungkin sudah tidak sinkron)
        cursor.execute("""
            SELECT k.nama_kelas,
                (SELECT COUNT(*) FROM pendaftaran p WHERE p.id_kelas = k.id_kelas) AS jumlah_siswa
            FROM kelas k
            WHERE k.id_kelas = %s
        """, (id_kelas,))
        row = cursor.fetchone()

        if not row:
            flash('Kelas tidak ditemukan.', 'error')
            return redirect(url_for('admin.manajemen_kelas_admin'))

        if row['jumlah_siswa'] > 0:
            flash('Kelas ini sudah memiliki siswa terdaftar sehingga tidak bisa dihapus.', 'error')
            return redirect(url_for('admin.edit_kelas', id_kelas=id_kelas))

        cursor.execute("DELETE FROM kelas WHERE id_kelas = %s", (id_kelas,))
        conn.commit()

        flash(f'Kelas "{row["nama_kelas"]}" berhasil dihapus.', 'success')
        return redirect(url_for('admin.manajemen_kelas_admin'))

    except Exception as e:
        conn.rollback()
        print(f"\n[ERROR HAPUS KELAS]: {e}\n")
        flash('Terjadi kegagalan sistem saat menghapus kelas.', 'error')
        return redirect(url_for('admin.edit_kelas', id_kelas=id_kelas))
    finally:
        cursor.close()
        conn.close()

@admin_bp.route('/manajemen_pengajar_admin')
def manajemen_pengajar_admin():
    # 1. Verifikasi hak akses admin login
    cek_akses = _wajib_login_admin()
    if cek_akses:
        return cek_akses
    
    # 2. Tangkap parameter filter, pencarian, dan halaman (Pagination)
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    per_page = 5  # Menampilkan 5 data per halaman
    offset = (page - 1) * per_page

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 3. Statistik: Hitung total pengajar aktif riil (Role R02)
        cursor.execute("SELECT COUNT(*) as total FROM users WHERE role_id_role = 'R02'")
        total_semua = cursor.fetchone()['total'] or 0

        # Sesi Kelas Aktif: Menghitung kelas riil yang statusnya 'Aktif'
        cursor.execute("SELECT COUNT(*) as total FROM kelas WHERE status_kelas = 'Aktif'")
        sesi_aktif = cursor.fetchone()['total'] or 0

        # Nilai representatif statis untuk penilaian & validasi pendaftaran
        rata_rating = "4.8"
        pending_validasi = 0

        # 4. Hitung total records pencarian untuk dasar pembuatan pagination halaman
        query_count = """
            SELECT COUNT(*) as total 
            FROM users 
            WHERE role_id_role = 'R02' AND (username LIKE %s OR email LIKE %s)
        """
        search_param = f"%{search}%"
        cursor.execute(query_count, (search_param, search_param))
        total_data = cursor.fetchone()['total'] or 0
        
        import math
        total_pages = max(1, math.ceil(total_data / per_page))

        # 5. AMBIL DATA RIIL: Diurutkan berdasarkan alfabet username (A-Z) karena tidak ada 'created_at'
        query_pengajar = """
            SELECT 
                u.id_users, 
                u.username, 
                u.email, 
                u.status_akun,
                (SELECT COUNT(*) FROM kelas k WHERE k.id_pengajar = u.id_users) AS total_kelas,
                (SELECT COUNT(DISTINCT p.id_anak) FROM pendaftaran p JOIN kelas k ON p.id_kelas = k.id_kelas WHERE k.id_pengajar = u.id_users) AS total_siswa
            FROM users u
            WHERE u.role_id_role = 'R02' AND (u.username LIKE %s OR u.email LIKE %s)
            ORDER BY u.username ASC
            LIMIT %s OFFSET %s
        """
        cursor.execute(query_pengajar, (search_param, search_param, per_page, offset))
        daftar_pengajar = cursor.fetchall()

        # 6. Kirim semua variabel ke file HTML
        return render_template(
            'admin/manajemen_pengajar.html',
            daftar_pengajar=daftar_pengajar,
            total_semua=total_semua,
            sesi_aktif=sesi_aktif,
            rata_rating=rata_rating,
            pending_validasi=pending_validasi,
            search=search,
            page=page,
            total_data=total_data,
            total_pages=total_pages
        )

    except Exception as e:
        # Mencetak pesan error asli ke console log terminal terminal/CMD
        print(f"\n[SISTEM EROR DATABASE PENGAJAR]: {e}\n")
        flash('Gagal memuat data pengajar secara riil.', 'error')
        
        return render_template(
            'admin/manajemen_pengajar.html',
            daftar_pengajar=[], total_semua=0, sesi_aktif=0, rata_rating="0.0", pending_validasi=0,
            search=search, page=1, total_data=0, total_pages=1
        )
    finally:
        cursor.close()
        conn.close()

@admin_bp.route('/admin/pengajar/detail/<id_pengajar>')
def detail_pengajar_admin(id_pengajar):
    # Proteksi Sesi Admin (Pastikan hanya Kepala / Admin yang bisa masuk)
    cek_akses = _wajib_login_admin()
    if cek_akses:
        return cek_akses
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 1. Query Data Utama Pengajar dari tabel users
        cursor.execute("""
            SELECT id_users, nama_lengkap, email, no_telp, foto_profil, status_akun 
            FROM users 
            WHERE id_users = %s AND role_id_role = 'R02'
        """, (id_pengajar,))
        pengajar = cursor.fetchone()
        
        if not pengajar:
            flash('Data pengajar tidak ditemukan!', 'error')
            return redirect(url_for('admin.dashboard_admin'))
            
        # 2. Query Daftar Kelas yang Ditugaskan ke Pengajar beserta kalkulasi jumlah siswa aktif saat ini
        cursor.execute("""
            SELECT 
                k.id_kelas, 
                k.nama_kelas, 
                k.tingkat, 
                k.kapasitas_maksimal AS kuota, 
                k.hari_jadwal AS hari, 
                k.harga,
                k.status_kelas,
                COUNT(p.id_pendaftaran) AS jumlah_siswa
            FROM kelas k
            LEFT JOIN pendaftaran p ON k.id_kelas = p.id_kelas AND p.status_pendaftaran = 'Aktif'
            WHERE k.id_pengajar = %s
            GROUP BY k.id_kelas
        """, (id_pengajar,))
        daftar_kelas = cursor.fetchall()

        # 2b. Hitung jumlah kelas yang masih aktif (masih diajar) -- dipakai untuk
        # mencegah admin membekukan akun pengajar yang masih punya kelas berjalan
        kelas_aktif_diajar = sum(1 for k in daftar_kelas if k.get('status_kelas') == 'Aktif')
        
        # 3. Hitung total seluruh siswa unik yang diajar oleh pengajar tersebut
        cursor.execute("""
            SELECT COUNT(DISTINCT p.id_anak) AS total_siswa
            FROM kelas k
            JOIN pendaftaran p ON k.id_kelas = p.id_kelas
            WHERE k.id_pengajar = %s AND p.status_pendaftaran = 'Aktif'
        """, (id_pengajar,))
        res_siswa = cursor.fetchone()
        total_siswa = res_siswa['total_siswa'] if res_siswa else 0
        
        # 4. Query Jadwal Sesi Kelas Terdekat dari tabel sesi_kelas
        cursor.execute("""
            SELECT sk.sesi_ke AS pertemuan_ke, sk.topik_pembahasan AS topik, 
                   k.nama_kelas, k.hari_jadwal AS hari, k.jam_mulai, sk.tanggal
            FROM sesi_kelas sk
            JOIN kelas k ON sk.id_kelas = k.id_kelas
            WHERE k.id_pengajar = %s AND sk.tanggal >= CURDATE()
            ORDER BY sk.tanggal ASC, k.jam_mulai ASC
            LIMIT 5
        """, (id_pengajar,))
        daftar_jadwal = cursor.fetchall()
        
    except Exception as e:
        print(f"Error pada server detail pengajar: {e}")
        flash('Terjadi kesalahan koneksi database.', 'error')
        return redirect(url_for('admin.dashboard_admin'))
    finally:
        cursor.close()
        conn.close()
        
    # Render ke halaman template baru dengan membawa variabel data dari DB
    return render_template('admin/detail_pengajar_admin.html', 
                           pengajar=pengajar, 
                           daftar_kelas=daftar_kelas, 
                           total_siswa=total_siswa, 
                           daftar_jadwal=daftar_jadwal,
                           kelas_aktif_diajar=kelas_aktif_diajar)

@admin_bp.route('/manajemen_orangtua_admin')
def manajemen_orangtua_admin():
    # Proteksi Sesi Admin (Kepala)
    cek_akses = _wajib_login_admin()
    if cek_akses:
        return cek_akses
    
    # 1. Tangkap Parameter Query String dari URL
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    per_page = 10  # Batas data per halaman table
    offset = (page - 1) * per_page
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # --- KELOMPOK STATISTIK DASHBOARD ---
        # A. Hitung Jumlah Keluarga (User dengan role 'murid'/orangtua)
        cursor.execute("""
            SELECT COUNT(*) as total_keluarga 
            FROM users u
            JOIN role r ON u.role_id_role = r.id_role
            WHERE r.nama_role = 'murid'
        """)
        total_keluarga = cursor.fetchone()['total_keluarga'] or 0
        
        # B. Hitung Siswa Aktif dari tabel anak
        cursor.execute("SELECT COUNT(*) as total_siswa_aktif FROM anak WHERE status_anak = 'Active'")
        total_siswa_aktif = cursor.fetchone()['total_siswa_aktif'] or 0
        
        # C. Hitung Pendaftaran Tertunda (Orang tua yang akunnya masih 'unverified')
        cursor.execute("""
            SELECT COUNT(*) as total_pending 
            FROM users u
            JOIN role r ON u.role_id_role = r.id_role
            WHERE r.nama_role = 'murid' AND u.status_akun = 'unverified'
        """)
        total_pending = cursor.fetchone()['total_pending'] or 0


        # --- QUERY DAFTAR ORANG TUA & JUMLAH ANAK ---
        # Menggunakan LEFT JOIN agar orang tua yang belum mendaftarkan anaknya tetap muncul di list admin
        base_query = """
            SELECT 
                u.id_users, 
                u.username, 
                u.email, 
                u.status_akun,
                COUNT(a.id_anak) as total_anak
            FROM users u
            JOIN role r ON u.role_id_role = r.id_role
            LEFT JOIN anak a ON u.id_users = a.id_orangtua
            WHERE r.nama_role = 'murid'
        """
        params = []
        
        # Filter Pencarian Berdasarkan Nama Pengguna atau Email
        if search:
            base_query += " AND (u.username LIKE %s OR u.email LIKE %s)"
            params.extend([f"%{search}%", f"%{search}%"])
            
        base_query += " GROUP BY u.id_users"
        
        # Hitung total baris data yang cocok (untuk kalkulasi halaman)
        cursor.execute(base_query, params)
        total_data = len(cursor.fetchall())
        total_pages = math.ceil(total_data / per_page)
        if total_pages == 0: 
            total_pages = 1
            
        # Eksekusi limitasi halaman data (Pagination)
        final_query = base_query + " ORDER BY u.username ASC LIMIT %s OFFSET %s"
        params.extend([per_page, offset])
        
        cursor.execute(final_query, params)
        daftar_orangtua = cursor.fetchall()
        
        # Kirim variabel data menuju context template Jinja2
        return render_template('admin/manajemen_orangtua_anak.html', 
                               daftar_orangtua=daftar_orangtua,
                               search=search,
                               page=page,
                               total_pages=total_pages,
                               total_data=total_data,
                               total_keluarga=total_keluarga,
                               total_siswa_aktif=total_siswa_aktif,
                               total_pending=total_pending)
                               
    except Exception as e:
        print(f"[ERROR] Gagal memuat manajemen orang tua & anak: {e}")
        flash('Terjadi kegagalan sistem saat mengambil data komunitas.', 'error')
        return render_template('admin/manajemen_orangtua_anak.html', 
                               daftar_orangtua=[], 
                               total_pages=1, 
                               page=1, 
                               total_data=0,
                               total_keluarga=0,
                               total_siswa_aktif=0,
                               total_pending=0)
    finally:
        cursor.close()
        conn.close()

@admin_bp.route('/manajemen_orangtua_admin/detail/<id_users>')
def detail_orangtua(id_users):
    cek_akses = _wajib_login_admin()
    if cek_akses:
        return cek_akses
        
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 1. Ambil Biodata Orang Tua
        cursor.execute("SELECT * FROM users WHERE id_users = %s AND role_id_role = 'R01'", (id_users,))
        orangtua = cursor.fetchone()
        
        if not orangtua:
            flash('Data wali murid tidak ditemukan.', 'error')
            return redirect(url_for('admin.manajemen_orangtua_admin'))
            
        # 2. Ambil Data Anak beserta Tingkat Sekolah & Kelas Bimbel yang diikuti
        cursor.execute("""
            SELECT a.id_anak, a.nama_lengkap, a.kelas as tingkat_sekolah, a.sekolah_asal,
                   GROUP_CONCAT(k.nama_kelas SEPARATOR ', ') as kelas_bimbel
            FROM anak a
            LEFT JOIN pendaftaran p ON a.id_anak = p.id_anak AND p.status_pendaftaran = 'Aktif'
            LEFT JOIN kelas k ON p.id_kelas = k.id_kelas
            WHERE a.id_orangtua = %s
            GROUP BY a.id_anak
        """, (id_users,))
        daftar_anak = cursor.fetchall()
        
        # 3. Ambil Riwayat Pembayaran/Transaksi SPP
        cursor.execute("""
            SELECT p.id_pembayaran, p.tanggal_bayar, p.jumlah_bayar, p.status_pembayaran,
                   a.nama_lengkap as nama_anak, k.nama_kelas
            FROM pembayaran p
            JOIN pendaftaran pd ON p.id_pendaftaran = pd.id_pendaftaran
            JOIN anak a ON pd.id_anak = a.id_anak
            JOIN kelas k ON pd.id_kelas = k.id_kelas
            WHERE a.id_orangtua = %s
            ORDER BY p.tanggal_bayar DESC
        """, (id_users,))
        riwayat_transaksi = cursor.fetchall()
        
        return render_template('admin/detail_orangtua.html', 
                               orangtua=orangtua, 
                               daftar_anak=daftar_anak, 
                               riwayat_transaksi=riwayat_transaksi)
    except Exception as e:
        print(f"[ERROR DETAIL ORANG TUA]: {e}")
        return redirect(url_for('admin.manajemen_orangtua_admin'))
    finally:
        cursor.close()
        conn.close()

@admin_bp.route('/manajemen_pengajar_admin/suspend/<id_users>', methods=['POST'])
def suspend_akun_pengajar(id_users):
    # Route untuk membekukan akun pengajar (mengubah status dari verified menjadi suspended)
    cek_akses = _wajib_login_admin()
    if cek_akses:
        return cek_akses

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Cek status saat ini
        cursor.execute("SELECT status_akun FROM users WHERE id_users = %s", (id_users,))
        current_status = cursor.fetchone()[0]

        # Toggle Status (Jika verified jadi suspended, jika suspended jadi verified)
        new_status = 'suspended' if current_status == 'verified' else 'verified'

        # Kalau yang mau dilakukan adalah MEMBEKUKAN (bukan mengaktifkan kembali),
        # cek ulang langsung dari database apakah pengajar ini masih punya kelas
        # berstatus 'Aktif' -- jangan percaya kondisi apapun dari sisi frontend,
        # karena tombol/modal di halaman bisa saja sudah tidak sinkron.
        if new_status == 'suspended':
            cursor.execute(
                "SELECT COUNT(*) FROM kelas WHERE id_pengajar = %s AND status_kelas = 'Aktif'",
                (id_users,)
            )
            jumlah_kelas_aktif = cursor.fetchone()[0] or 0

            if jumlah_kelas_aktif > 0:
                flash(
                    f'Akun tidak bisa dibekukan karena pengajar ini masih memiliki '
                    f'{jumlah_kelas_aktif} kelas aktif yang sedang diajar. '
                    f'Pindahkan atau selesaikan kelas tersebut terlebih dahulu.',
                    'error'
                )
                return redirect(url_for('admin.detail_pengajar_admin', id_pengajar=id_users))

        cursor.execute("UPDATE users SET status_akun = %s WHERE id_users = %s", (new_status, id_users))
        conn.commit()

        status_msg = 'dibekukan' if new_status == 'suspended' else 'diaktifkan kembali'
        flash(f'Akun pengajar berhasil {status_msg}.', 'success')

    except Exception as e:
        conn.rollback()
        print(f"[ERROR SUSPEND AKUN PENGAJAR]: {e}")
        flash('Gagal mengubah status akun pengajar.', 'error')
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('admin.detail_pengajar_admin', id_pengajar=id_users))

@admin_bp.route('/tambah_pengajar', methods=['POST'])
def tambah_pengajar():
    # Pastikan yang mengakses adalah Admin (Kepala)
    cek_akses = _wajib_login_admin()
    if cek_akses:
        return cek_akses
    
    username = request.form.get('username')
    email = request.form.get('email')
    password = request.form.get('password')
    
    # Validasi input kosong
    if not username or not email or not password:
        flash('Semua data wajib diisi!', 'error')
        return redirect(url_for('admin.manajemen_pengajar_admin'))
        
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 1. Cek apakah email sudah terdaftar sebelumnya
        cursor.execute("SELECT id_users FROM users WHERE email = %s", (email,))
        if cursor.fetchone():
            flash('Email sudah terdaftar!', 'error')
            return redirect(url_for('admin.manajemen_pengajar_admin'))
        
        # 2. Ambil id_role untuk 'Pengajar' dari tabel role secara dinamis
        cursor.execute("SELECT id_role FROM role WHERE nama_role = 'Pengajar'")
        role_data = cursor.fetchone()
        if not role_data:
            flash('Role Pengajar tidak ditemukan di database!', 'error')
            return redirect(url_for('admin.manajemen_pengajar_admin'))
            
        role_id = role_data['id_role']
        
        # 3. Generate ID unik untuk pengajar baru (Format UUID 10 karakter seperti sistem OTP Anda)
        new_user_id = str(uuid.uuid4().hex[:10]).upper()
        
        # 4. Insert data ke tabel users
        # Catatan: Password disimpan dalam bentuk plain text agar sesuai dengan fungsi 
        # `login_pengajar_action` Anda yang menggunakan perbandingan langsung (`== password`)
        cursor.execute("""
            INSERT INTO users (id_users, username, email, password, role_id_role, status_akun)
            VALUES (%s, %s, %s, %s, %s, 'verified')
        """, (new_user_id, username, email, password, role_id))
        
        conn.commit()
        flash('Pengajar baru berhasil ditambahkan!', 'success')
        
    except Exception as e:
        conn.rollback()
        print(f"Error saat menambah pengajar: {e}")
        flash('Terjadi kesalahan pada server saat menambahkan pengajar.', 'error')
    finally:
        cursor.close()
        conn.close()
        
    return redirect(url_for('admin.manajemen_pengajar_admin'))


# ==========================================================
#                VALIDASI PEMBAYARAN (ADMIN)
# ==========================================================

@admin_bp.route('/validasi_pembayaran_admin')
def validasi_pembayaran_admin():
    # 1. Proteksi Sesi Admin (Kepala)
    cek_akses = _wajib_login_admin()
    if cek_akses:
        return cek_akses

    # 2. Tangkap parameter pencarian, filter status, dan halaman untuk tabel riwayat
    search = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    page = request.args.get('page', 1, type=int)
    per_page = 8
    offset = (page - 1) * per_page

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # --- KELOMPOK STATISTIK KARTU ATAS ---
        # A. Tugas yang tertunda (menunggu persetujuan)
        cursor.execute("SELECT COUNT(*) as total FROM pembayaran WHERE status_pembayaran = 'Pending'")
        tugas_tertunda = cursor.fetchone()['total'] or 0

        # B. Divalidasi hari ini (disetujui/ditolak hari ini)
        cursor.execute("""
            SELECT COUNT(*) as total FROM pembayaran 
            WHERE status_pembayaran IN ('Lunas', 'Ditolak') AND DATE(tanggal_bayar) = CURDATE()
        """)
        divalidasi_hari_ini = cursor.fetchone()['total'] or 0

        # C. Total nominal yang sedang menunggu validasi
        cursor.execute("SELECT COALESCE(SUM(jumlah_bayar), 0) as total FROM pembayaran WHERE status_pembayaran = 'Pending'")
        total_nominal_pending = cursor.fetchone()['total'] or 0

        # D. Perlu perhatian: pengajuan pending TANPA bukti bayar, atau sudah
        # menunggu (belum diurus) selama 1 hari lebih sejak diajukan.
        cursor.execute("""
            SELECT COUNT(*) as total FROM pembayaran
            WHERE status_pembayaran = 'Pending'
              AND (bukti_bayar IS NULL OR bukti_bayar = '' OR tanggal_bayar < (NOW() - INTERVAL 1 DAY))
        """)
        perlu_perhatian = cursor.fetchone()['total'] or 0

        # --- ANTREAN: MENUNGGU PERSETUJUAN (ditampilkan di bagian atas halaman) ---
        cursor.execute("""
            SELECT 
                p.id_pembayaran,
                CONCAT('TXN-', LPAD(p.id_pembayaran, 5, '0')) AS kode_transaksi,
                p.jumlah_bayar,
                p.tanggal_bayar,
                p.status_pembayaran,
                p.bukti_bayar,
                p.keterangan,
                a.nama_lengkap AS nama_anak,
                k.nama_kelas,
                u.username AS nama_orangtua
            FROM pembayaran p
            JOIN pendaftaran pd ON p.id_pendaftaran = pd.id_pendaftaran
            JOIN anak a ON pd.id_anak = a.id_anak
            JOIN kelas k ON pd.id_kelas = k.id_kelas
            JOIN users u ON a.id_orangtua = u.id_users
            WHERE p.status_pembayaran = 'Pending'
            ORDER BY p.tanggal_bayar ASC
        """)
        antrean_pembayaran = cursor.fetchall()
        batas_perlu_perhatian = datetime.now() - timedelta(days=1)
        for item in antrean_pembayaran:
            item['tanggal_bayar_fmt'] = format_tanggal_indo(item['tanggal_bayar'])
            # Tandai item yang "perlu perhatian": belum ada bukti bayar, atau
            # sudah menunggu (belum diurus/divalidasi) selama 1 hari lebih.
            item['perlu_perhatian'] = (
                not item.get('bukti_bayar')
                or (item.get('tanggal_bayar') and item['tanggal_bayar'] < batas_perlu_perhatian)
            )

        # --- TABEL RIWAYAT SEMUA TRANSAKSI (dengan pencarian, filter status & pagination) ---
        base_query = """
            FROM pembayaran p
            JOIN pendaftaran pd ON p.id_pendaftaran = pd.id_pendaftaran
            JOIN anak a ON pd.id_anak = a.id_anak
            JOIN kelas k ON pd.id_kelas = k.id_kelas
            JOIN users u ON a.id_orangtua = u.id_users
            WHERE (a.nama_lengkap LIKE %s OR u.username LIKE %s OR k.nama_kelas LIKE %s)
        """
        search_param = f"%{search}%"
        params = [search_param, search_param, search_param]

        if status_filter in ('Pending', 'Lunas', 'Ditolak'):
            base_query += " AND p.status_pembayaran = %s"
            params.append(status_filter)

        cursor.execute(f"SELECT COUNT(*) as total {base_query}", params)
        total_data = cursor.fetchone()['total'] or 0
        total_pages = max(1, math.ceil(total_data / per_page))

        cursor.execute(f"""
            SELECT 
                p.id_pembayaran,
                CONCAT('TXN-', LPAD(p.id_pembayaran, 5, '0')) AS kode_transaksi,
                p.jumlah_bayar,
                p.tanggal_bayar,
                p.status_pembayaran,
                p.bukti_bayar,
                p.keterangan,
                a.nama_lengkap AS nama_anak,
                k.nama_kelas,
                u.username AS nama_orangtua
            {base_query}
            ORDER BY p.tanggal_bayar DESC
            LIMIT %s OFFSET %s
        """, params + [per_page, offset])
        semua_transaksi = cursor.fetchall()
        for t in semua_transaksi:
            t['tanggal_bayar_fmt'] = format_tanggal_indo(t['tanggal_bayar'])

        return render_template(
            'admin/validasi_pembayaran.html',
            tugas_tertunda=tugas_tertunda,
            divalidasi_hari_ini=divalidasi_hari_ini,
            total_nominal_pending=total_nominal_pending,
            perlu_perhatian=perlu_perhatian,
            antrean_pembayaran=antrean_pembayaran,
            semua_transaksi=semua_transaksi,
            search=search,
            status_filter=status_filter,
            page=page,
            total_pages=total_pages,
            total_data=total_data
        )

    except Exception as e:
        print(f"\n[ERROR VALIDASI PEMBAYARAN]: {e}\n")
        flash('Gagal memuat data pembayaran dari database.', 'error')
        return render_template(
            'admin/validasi_pembayaran.html',
            tugas_tertunda=0, divalidasi_hari_ini=0, total_nominal_pending=0, perlu_perhatian=0,
            antrean_pembayaran=[], semua_transaksi=[], search=search, status_filter=status_filter,
            page=1, total_pages=1, total_data=0
        )
    finally:
        cursor.close()
        conn.close()


@admin_bp.route('/validasi_pembayaran_admin/download_laporan')
def download_laporan_pembayaran():
    # Proteksi Sesi Admin (Kepala)
    cek_akses = _wajib_login_admin()
    if cek_akses:
        return cek_akses

    # Mengikuti filter pencarian & status yang sedang aktif di halaman,
    # tapi TANPA pagination, supaya laporan berisi seluruh data yang sesuai filter.
    search = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    periode = request.args.get('periode', '')  # '', 'hari_ini', '7_hari', '1_bulan'

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        base_query = """
            FROM pembayaran p
            JOIN pendaftaran pd ON p.id_pendaftaran = pd.id_pendaftaran
            JOIN anak a ON pd.id_anak = a.id_anak
            JOIN kelas k ON pd.id_kelas = k.id_kelas
            JOIN users u ON a.id_orangtua = u.id_users
            WHERE (a.nama_lengkap LIKE %s OR u.username LIKE %s OR k.nama_kelas LIKE %s)
        """
        search_param = f"%{search}%"
        params = [search_param, search_param, search_param]

        if status_filter in ('Pending', 'Lunas', 'Ditolak'):
            base_query += " AND p.status_pembayaran = %s"
            params.append(status_filter)

        # --- Filter rentang tanggal (periode laporan) ---
        # PENTING: rentang tanggal dihitung di Python (bukan CURDATE()/NOW() milik
        # MySQL) supaya hasilnya tidak bergantung pada timezone server database.
        PERIODE_LABEL = {
            'hari_ini': 'Hari Ini',
            'minggu_ini': 'Minggu Ini',
            'bulan_ini': 'Bulan Ini',
        }
        if periode in PERIODE_LABEL:
            sekarang = datetime.now()
            awal_hari_ini = sekarang.replace(hour=0, minute=0, second=0, microsecond=0)
            batas_akhir = awal_hari_ini + timedelta(days=1)  # eksklusif, mencakup s.d. akhir hari ini

            if periode == 'hari_ini':
                batas_awal = awal_hari_ini
            elif periode == 'minggu_ini':
                # Senin adalah awal minggu (weekday(): Senin=0 ... Minggu=6)
                batas_awal = awal_hari_ini - timedelta(days=awal_hari_ini.weekday())
            elif periode == 'bulan_ini':
                batas_awal = awal_hari_ini.replace(day=1)

            base_query += " AND p.tanggal_bayar >= %s AND p.tanggal_bayar < %s"
            params.append(batas_awal)
            params.append(batas_akhir)

        cursor.execute(f"""
            SELECT
                CONCAT('TXN-', LPAD(p.id_pembayaran, 5, '0')) AS kode_transaksi,
                a.nama_lengkap AS nama_anak,
                u.username AS nama_orangtua,
                k.nama_kelas,
                p.tanggal_bayar,
                p.jumlah_bayar,
                p.status_pembayaran,
                p.keterangan
            {base_query}
            ORDER BY p.tanggal_bayar DESC
        """, params)
        rows = cursor.fetchall()

        # --- Membuat file Excel (.xlsx) yang rapi di memory, langsung dikirim
        # sebagai file download tanpa perlu disimpan fisik di server ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Pembayaran"

        WARNA_UTAMA = "A22754"
        font_umum = "Calibri"

        headers = [
            'Kode Transaksi', 'Nama Siswa', 'Nama Orang Tua', 'Kelas',
            'Tanggal Bayar', 'Jumlah Bayar (Rp)', 'Status', 'Keterangan'
        ]

        # --- Judul laporan di baris paling atas ---
        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        judul_cell = ws['A1']
        judul_cell.value = "LAPORAN PEMBAYARAN - TEC ENGLISH COURSE"
        judul_cell.font = Font(name=font_umum, size=14, bold=True, color=WARNA_UTAMA)
        judul_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 26

        ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
        sub_cell = ws['A2']
        filter_info = []
        if periode in PERIODE_LABEL:
            filter_info.append(f"Periode: {PERIODE_LABEL[periode]}")
        if search:
            filter_info.append(f"Pencarian: \"{search}\"")
        if status_filter:
            filter_info.append(f"Status: {status_filter}")
        keterangan_filter = " | ".join(filter_info) if filter_info else "Semua Data"
        sub_cell.value = f"Dicetak: {datetime.now().strftime('%d %B %Y, %H:%M')} WIB   |   {keterangan_filter}   |   Total: {len(rows)} transaksi"
        sub_cell.font = Font(name=font_umum, size=9, italic=True, color="808080")
        sub_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 18

        # --- Baris header tabel (baris ke-4) ---
        header_row = 4
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )
        for col_idx, judul in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=judul)
            cell.font = Font(name=font_umum, size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=WARNA_UTAMA, end_color=WARNA_UTAMA, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 22

        # --- Isi data ---
        status_fill_map = {
            'Lunas':   PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
            'Pending': PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            'Ditolak': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        }
        status_font_map = {
            'Lunas':   Font(name=font_umum, size=10, bold=True, color="16A34A"),
            'Pending': Font(name=font_umum, size=10, bold=True, color="D97706"),
            'Ditolak': Font(name=font_umum, size=10, bold=True, color="DC2626"),
        }

        for i, r in enumerate(rows):
            row_num = header_row + 1 + i
            nilai_baris = [
                r['kode_transaksi'],
                r['nama_anak'],
                r['nama_orangtua'],
                r['nama_kelas'],
                format_tanggal_indo(r['tanggal_bayar']),
                r['jumlah_bayar'],
                r['status_pembayaran'],
                r['keterangan'] or '-'
            ]
            zebra_fill = PatternFill(start_color="FFF9FA", end_color="FFF9FA", fill_type="solid") if i % 2 else None

            for col_idx, nilai in enumerate(nilai_baris, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=nilai)
                cell.font = Font(name=font_umum, size=10)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=(col_idx == 8))

                if col_idx == 6:  # kolom Jumlah Bayar -> format rupiah
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col_idx == 7:  # kolom Status -> highlight warna
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if nilai in status_fill_map:
                        cell.fill = status_fill_map[nilai]
                        cell.font = status_font_map[nilai]
                    continue  # jangan ditimpa zebra_fill di bawah
                elif col_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                if zebra_fill and col_idx != 7:
                    cell.fill = zebra_fill

        # --- Baris total di akhir tabel ---
        total_row = header_row + 1 + len(rows)
        ws.merge_cells(f'A{total_row}:E{total_row}')
        total_label = ws.cell(row=total_row, column=1, value="TOTAL")
        total_label.font = Font(name=font_umum, size=10, bold=True)
        total_label.alignment = Alignment(horizontal='right', vertical='center')
        total_label.border = thin_border
        for col in range(2, 6):
            ws.cell(row=total_row, column=col).border = thin_border

        total_nominal_cell = ws.cell(row=total_row, column=6, value=sum(r['jumlah_bayar'] for r in rows))
        total_nominal_cell.font = Font(name=font_umum, size=10, bold=True, color=WARNA_UTAMA)
        total_nominal_cell.number_format = '#,##0'
        total_nominal_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_nominal_cell.fill = PatternFill(start_color="FFF0F3", end_color="FFF0F3", fill_type="solid")
        total_nominal_cell.border = thin_border
        for col in (7, 8):
            c = ws.cell(row=total_row, column=col)
            c.border = thin_border
            c.fill = PatternFill(start_color="FFF0F3", end_color="FFF0F3", fill_type="solid")

        # --- Lebar kolom & pengaturan lain ---
        lebar_kolom = [16, 22, 20, 22, 20, 18, 12, 28]
        for i, lebar in enumerate(lebar_kolom, start=1):
            ws.column_dimensions[get_column_letter(i)].width = lebar

        ws.freeze_panes = f"A{header_row + 1}"
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row}"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        suffix_periode = f"_{periode}" if periode in PERIODE_LABEL else ""
        nama_file = f"laporan_pembayaran{suffix_periode}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={nama_file}'}
        )

    except Exception as e:
        print(f"\n[ERROR DOWNLOAD LAPORAN PEMBAYARAN]: {e}\n")
        flash('Gagal membuat laporan pembayaran.', 'error')
        return redirect(url_for('admin.validasi_pembayaran_admin'))
    finally:
        cursor.close()
        conn.close()


@admin_bp.route('/validasi_pembayaran_admin/setujui/<int:id_pembayaran>', methods=['POST'])
def setujui_pembayaran(id_pembayaran):
    # Proteksi Sesi Admin (Kepala)
    cek_akses = _wajib_login_admin()
    if cek_akses:
        return cek_akses

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # Ambil data yang dibutuhkan untuk notifikasi email SEBELUM di-update,
        # supaya tetap bisa dipakai meski setelah update baris pembayarannya
        # tidak lagi berstatus 'Pending'.
        cursor.execute("""
            SELECT 
                u.id_users AS id_orangtua,
                u.email AS email_orangtua,
                u.username AS nama_orangtua,
                u.notif_tagihan,
                a.nama_lengkap AS nama_anak,
                k.nama_kelas,
                pb.jumlah_bayar,
                pb.kode_pembayaran
            FROM pembayaran pb
            JOIN pendaftaran pd ON pd.id_pendaftaran = pb.id_pendaftaran
            JOIN anak a ON a.id_anak = pd.id_anak
            JOIN users u ON u.id_users = a.id_orangtua
            JOIN kelas k ON k.id_kelas = pd.id_kelas
            WHERE pb.id_pembayaran = %s
        """, (id_pembayaran,))
        info = cursor.fetchone()

        # Hanya memproses transaksi yang statusnya masih 'Pending' (mencegah proses ganda).
        # Sekalian aktifkan pendaftaran muridnya (status_pendaftaran -> 'Aktif'), karena
        # sebelum ini murid baru berstatus 'Pending' menunggu verifikasi pembayaran.
        cursor.execute("""
            UPDATE pembayaran pb
            JOIN pendaftaran pd ON pd.id_pendaftaran = pb.id_pendaftaran
            SET pb.status_pembayaran = 'Lunas',
                pd.status_pendaftaran = 'Aktif'
            WHERE pb.id_pembayaran = %s AND pb.status_pembayaran = 'Pending'
        """, (id_pembayaran,))
        conn.commit()

        if cursor.rowcount > 0:
            flash('Pembayaran berhasil disetujui dan ditandai Lunas.', 'success')

            # Kirim notifikasi email ke orang tua, kecuali mereka mematikan
            # notifikasi tagihan di pengaturan akun (notif_tagihan = 0)
            if info and info.get('email_orangtua') and info.get('notif_tagihan'):
                try:
                    send_email(
                        to=info['email_orangtua'],
                        subject='Pembayaran Anda Telah Diverifikasi - TEC Portal',
                        body=(
                            f"Halo {info['nama_orangtua']},\n\n"
                            f"Pembayaran Anda untuk kelas \"{info['nama_kelas']}\" atas nama {info['nama_anak']} "
                            f"(kode: {info.get('kode_pembayaran') or '-'}, jumlah: Rp {info['jumlah_bayar']:,}) "
                            f"telah diverifikasi dan dinyatakan LUNAS oleh admin.\n\n"
                            f"{info['nama_anak']} sekarang sudah aktif dan bisa mengikuti kelas.\n\n"
                            f"Terima kasih,\nTEC Portal"
                        )
                    )
                except Exception as email_err:
                    print(f"[ERROR] Gagal mengirim email notifikasi setujui pembayaran: {email_err}")
        else:
            flash('Pembayaran tidak ditemukan atau sudah diproses sebelumnya.', 'error')

    except Exception as e:
        conn.rollback()
        print(f"\n[ERROR SETUJUI PEMBAYARAN]: {e}\n")
        flash('Terjadi kesalahan pada server saat menyetujui pembayaran.', 'error')
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('admin.validasi_pembayaran_admin'))


@admin_bp.route('/validasi_pembayaran_admin/tolak/<int:id_pembayaran>', methods=['POST'])
def tolak_pembayaran(id_pembayaran):
    # Proteksi Sesi Admin (Kepala)
    cek_akses = _wajib_login_admin()
    if cek_akses:
        return cek_akses

    alasan = request.form.get('alasan_tolak', '').strip()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # Ambil data yang dibutuhkan untuk notifikasi email SEBELUM di-update,
        # supaya tetap bisa dipakai meski setelah update baris pembayarannya
        # tidak lagi berstatus 'Pending'.
        cursor.execute("""
            SELECT 
                u.id_users AS id_orangtua,
                u.email AS email_orangtua,
                u.username AS nama_orangtua,
                u.notif_tagihan,
                a.nama_lengkap AS nama_anak,
                k.nama_kelas,
                pb.jumlah_bayar,
                pb.kode_pembayaran
            FROM pembayaran pb
            JOIN pendaftaran pd ON pd.id_pendaftaran = pb.id_pendaftaran
            JOIN anak a ON a.id_anak = pd.id_anak
            JOIN users u ON u.id_users = a.id_orangtua
            JOIN kelas k ON k.id_kelas = pd.id_kelas
            WHERE pb.id_pembayaran = %s
        """, (id_pembayaran,))
        info = cursor.fetchone()

        # Sekalian tandai pendaftaran muridnya sebagai 'Ditolak', supaya murid yang
        # pembayarannya ditolak tidak nyangkut selamanya di status 'Pending'
        cursor.execute("""
            UPDATE pembayaran pb
            JOIN pendaftaran pd ON pd.id_pendaftaran = pb.id_pendaftaran
            SET pb.status_pembayaran = 'Ditolak',
                pb.keterangan = %s,
                pd.status_pendaftaran = 'Ditolak'
            WHERE pb.id_pembayaran = %s AND pb.status_pembayaran = 'Pending'
        """, (alasan if alasan else 'Ditolak oleh admin', id_pembayaran))
        conn.commit()

        if cursor.rowcount > 0:
            flash('Pembayaran berhasil ditolak.', 'success')

            # Kirim notifikasi email ke orang tua, kecuali mereka mematikan
            # notifikasi tagihan di pengaturan akun (notif_tagihan = 0)
            if info and info.get('email_orangtua') and info.get('notif_tagihan'):
                try:
                    send_email(
                        to=info['email_orangtua'],
                        subject='Pembayaran Anda Ditolak - TEC Portal',
                        body=(
                            f"Halo {info['nama_orangtua']},\n\n"
                            f"Mohon maaf, pembayaran Anda untuk kelas \"{info['nama_kelas']}\" atas nama {info['nama_anak']} "
                            f"(kode: {info.get('kode_pembayaran') or '-'}, jumlah: Rp {info['jumlah_bayar']:,}) "
                            f"DITOLAK oleh admin.\n\n"
                            f"Alasan: {alasan if alasan else 'Tidak disebutkan'}\n\n"
                            f"Silakan login ke TEC Portal untuk mengunggah ulang bukti pembayaran yang valid.\n\n"
                            f"Terima kasih,\nTEC Portal"
                        )
                    )
                except Exception as email_err:
                    print(f"[ERROR] Gagal mengirim email notifikasi tolak pembayaran: {email_err}")
        else:
            flash('Pembayaran tidak ditemukan atau sudah diproses sebelumnya.', 'error')

    except Exception as e:
        conn.rollback()
        print(f"\n[ERROR TOLAK PEMBAYARAN]: {e}\n")
        flash('Terjadi kesalahan pada server saat menolak pembayaran.', 'error')
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('admin.validasi_pembayaran_admin'))